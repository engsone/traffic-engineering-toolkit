# -*- coding: utf-8 -*-
"""
يشغل مولّد بطاقة أداء السلامة ويصدر PPTX + PDF باسم واضح:
بطاقة_اداء_السلامة_بمنطقة_<المنطقة>_لشهر_<الشهر>_<السنة>.pptx/pdf
"""
from pathlib import Path
import re
import sys
import datetime as dt
import win32com.client
import pythoncom
from openpyxl import load_workbook

BASE = Path(r'C:/Users/Lenovo-P51/Downloads/ppt_achievement_inspect')
TEMPLATE = BASE / 'source.pptx'
DATA = Path(r'C:/Users/Lenovo-P51/OneDrive/Desktop/بطاقة الاداء/شيت تحديث بطاقة الاداء.xlsx')
OUT_DIR = Path(r'C:/Users/Lenovo-P51/OneDrive/Desktop/بطاقة الاداء')

sys.path.insert(0, str(BASE))
from generate_safety_card import generate  # noqa: E402


def clean_part(s: str) -> str:
    s = str(s or '').strip()
    s = re.sub(r'[\\/:*?"<>|]+', '', s)
    s = re.sub(r'\s+', '_', s)
    return s


def read_general():
    wb = load_workbook(DATA, data_only=True)
    ws = wb['01_بيانات_عامة'] if '01_بيانات_عامة' in wb.sheetnames else wb['بيانات_عامة']
    values = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        values[str(row[0]).strip()] = row[1]
    region = values.get('المنطقة', 'تبوك')
    month = values.get('الشهر', 'يونيو')
    date_value = values.get('التاريخ', '')
    year = extract_year(date_value)
    return str(region).strip(), str(month).strip(), str(year).strip()


def extract_year(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.year
    s = str(v or '').strip()
    # supports 25/06/2026, 2026-06-25, etc.
    m = re.search(r'(20\d{2}|14\d{2})', s)
    if m:
        return m.group(1)
    return dt.datetime.now().year


def export_pdf(pptx_path: Path, pdf_path: Path):
    app = None
    pythoncom.CoInitialize()
    try:
        app = win32com.client.Dispatch('PowerPoint.Application')
        app.Visible = 1
        pres = app.Presentations.Open(str(pptx_path.resolve()), WithWindow=False)
        try:
            pres.SaveAs(str(pdf_path.resolve()), 32)  # PDF
        finally:
            pres.Close()
    finally:
        if app:
            app.Quit()


def main():
    region, month, year = read_general()
    name = f"بطاقة_اداء_السلامة_بمنطقة_{clean_part(region)}_لشهر_{clean_part(month)}_{clean_part(year)}"
    pptx_path = OUT_DIR / f'{name}.pptx'
    pdf_path = OUT_DIR / f'{name}.pdf'

    logs = generate(TEMPLATE, DATA, pptx_path)
    export_pdf(pptx_path, pdf_path)

    print('DONE')
    print(f'PPTX={pptx_path}')
    print(f'PDF={pdf_path}')
    for log in logs:
        print(log)


if __name__ == '__main__':
    main()
