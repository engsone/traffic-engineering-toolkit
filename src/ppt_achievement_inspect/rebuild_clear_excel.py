from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import shutil

old=Path(r'C:/Users/Lenovo-P51/OneDrive/Desktop/بطاقة الاداء/شيت تحديث بطاقة الاداء.xlsx')
backup=old.with_name('شيت تحديث بطاقة الاداء - نسخة قبل التوضيح.xlsx')
if old.exists() and not backup.exists():
    shutil.copyfile(old, backup)

vals={}; texts=[]
if old.exists():
    wb_old=load_workbook(old, data_only=True)
    if 'بيانات_عامة' in wb_old.sheetnames:
        for r in wb_old['بيانات_عامة'].iter_rows(min_row=2, values_only=True):
            vals[('general',str(r[0]).strip())]=r[1]
    if 'مؤشرات_صفحة_1' in wb_old.sheetnames:
        for r in wb_old['مؤشرات_صفحة_1'].iter_rows(min_row=2, values_only=True):
            vals[(str(r[0]).strip(),str(r[1]).strip())]=r[2]
    if 'رسوم_صفحة_1' in wb_old.sheetnames:
        for r in wb_old['رسوم_صفحة_1'].iter_rows(min_row=2, values_only=True):
            vals[(str(r[0]).strip(),str(r[1]).strip())]=r[2]
    if 'نصوص_صفحة_2' in wb_old.sheetnames:
        for r in wb_old['نصوص_صفحة_2'].iter_rows(min_row=2, values_only=True):
            texts.append(r)

def gv(k,d=''):
    v=vals.get(('general',k),d); return d if v is None else v

def kv(sec,ind,d=''):
    v=vals.get((sec,ind),d); return d if v is None else v

def cv(chart,cat,d=''):
    v=vals.get((chart,cat),d); return d if v is None else v

wb=Workbook(); ws=wb.active; ws.title='00_اقرأني'
blue='407189'; pale='FFF8E1'
thin=Side(style='thin', color='D7D7D7'); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def setup(ws, headers, widths=None):
    ws.sheet_view.rightToLeft=True; ws.freeze_panes='A2'
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h)
        cell.fill=PatternFill('solid', fgColor=blue)
        cell.font=Font(color='FFFFFF', bold=True)
        cell.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border=border
        ws.column_dimensions[get_column_letter(c)].width=(widths[c-1] if widths and c<=len(widths) else max(16,min(40,len(str(h))+8)))
    ws.auto_filter.ref=f'A1:{get_column_letter(len(headers))}1'

def style_data(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border=border
            cell.alignment=Alignment(horizontal='right', vertical='top', wrap_text=True)

def append_rows(ws, rows):
    for r in rows: ws.append(r)
    style_data(ws)

setup(ws,['مهم','التوضيح'],[28,90])
append_rows(ws,[
('طريقة الاستخدام','عدّل فقط الخانات الصفراء في كل تبويب، ثم احفظ الملف.'),
('التحويلات','أدخل عدد التحويلات النظامية وعدد التحويلات عالية الخطورة. المجموع يحسب تلقائيًا ويظهر في المربع الكبير أعلى رسم التحويلات.'),
('حالة التعداد المروري','أدخل عدد الوحدات التي تعمل وعدد الوحدات التي تحتاج صيانة. الرسم يأخذ الخيارين فقط.'),
('لا تغيّر','لا تغيّر أسماء التبويبات أو العناوين الزرقاء لأنها مربوطة بالمولّد.'),
])

ws=wb.create_sheet('01_بيانات_عامة'); setup(ws,['الحقل','القيمة','أين يظهر؟','ملاحظات'],[28,28,34,50])
append_rows(ws,[
('المنطقة',gv('المنطقة','تبوك'),'العنوان أعلى الشرائح','مثال: تبوك'),
('الشهر',gv('الشهر','يونيو'),'أعلى الشرائح','بدون أقواس'),
('التاريخ',gv('التاريخ','03/07/2025'),'أعلى الشرائح','صيغة يوم/شهر/سنة'),
('كبير المهندسين',gv('كبير المهندسين','م حســـان الــحلــبي'),'أعلى الشرائح',''),
('ملاحظة أسفل صفحة 1',gv('صفحة 1 - ملاحظات سفلية','التفاصيل الإضافية-يتم ذكرها في الصفحة التالية مع الترميز (*,**,***)'),'أسفل الشريحة الأولى',''),
])

ws=wb.create_sheet('02_التحويلات'); setup(ws,['البند','العدد','هل يدخل في المجموع؟','أين يظهر؟','ملاحظة'],[34,16,22,40,60])
append_rows(ws,[
('عدد التحويلات النظامية',kv('التحويلات','عدد التحويلات النظامية',0),'نعم','الرسم + المجموع الكبير','أدخل عدد التحويلات النظامية فقط'),
('عدد التحويلات عالية الخطورة',kv('التحويلات','عدد التحويلات عالية الخطورة',0),'نعم','الرسم + المجموع الكبير','أدخل عدد التحويلات عالية الخطورة فقط'),
('إجمالي التحويلات','=SUM(B2:B3)','محسوب تلقائيًا','المربع الذهبي الكبير فوق الرسم','لا تعدّل هذه الخانة'),
])
for cell in ['B2','B3']:
    ws[cell].fill=PatternFill('solid', fgColor=pale); ws[cell].comment=Comment('عدّل هذه الخانة فقط', 'Hermes')

ws=wb.create_sheet('03_SOR'); setup(ws,['البند','العدد','أين يظهر؟','ملاحظة'],[38,16,44,60])
append_rows(ws,[
('إجمالي الملاحظات المرصودة سابقًا',kv('SOR','المرصودة سابقا',0),'رقم المرصودة سابقا + رسم SOR',''),
('ملاحظات مرصودة هذا الشهر',kv('SOR','المرصودة هذا الشهر',0),'رقم هذا الشهر + رسم SOR',''),
('مغلق هذا الشهر',cv('عدد ملاحظات السلامة المرصودة SOR','مغلق  (هذا الشهر)',0),'رسم SOR',''),
('رفع أمر هذا الشهر',cv('عدد ملاحظات السلامة المرصودة SOR','رفع امر  (هذا الشهر)',0),'رسم SOR',''),
])
for r in range(2,6): ws[f'B{r}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('04_الحوادث'); setup(ws,['البند','العدد','أين يظهر؟','ملاحظة'],[34,16,44,60])
append_rows(ws,[('الإصابات',cv('حوادث الوفيات والإصابات','اصابات',0),'رسم حوادث الوفيات والإصابات',''),('الوفيات',cv('حوادث الوفيات والإصابات','وفيات',0),'رسم حوادث الوفيات والإصابات','')])
for r in range(2,4): ws[f'B{r}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('05_أوامر_العمل'); setup(ws,['التصنيف','العدد','أين يظهر؟','ملاحظة'],[42,16,44,60])
append_rows(ws,[
('موافقة A',cv('أوامر العمل','موافقة A',0),'رسم أوامر العمل',''),
('موافقة مع وجود ملاحظات B',cv('أوامر العمل','موافقة مع وجود ملاحظات B',0),'رسم أوامر العمل',''),
('إعادة تسليم C',cv('أوامر العمل','إعادة تسليم  C',0),'رسم أوامر العمل',''),
('رفض D',cv('أوامر العمل','رفض D',0),'رسم أوامر العمل',''),
('عدم ممانعه E',cv('أوامر العمل','عدم ممانعه E',0),'رسم أوامر العمل',''),
('إجمالي أوامر العمل التي تم مراجعتها','=SUM(B2:B6)','المربع الذهبي/العنوان','محسوب تلقائيًا'),
])
for r in range(2,7): ws[f'B{r}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('06_التعداد_المروري'); setup(ws,['حالة وحدة التعداد','العدد','أين يظهر؟','ملاحظة'],[34,16,44,60])
append_rows(ws,[
('تعمل',cv('حالة وحدات التعداد المروري','تعمل',4),'رسم حالة وحدات التعداد المروري','الوحدات العاملة'),
('تحتاج صيانة',cv('حالة وحدات التعداد المروري','تتطلب الى صيانة',1),'رسم حالة وحدات التعداد المروري','الوحدات التي تحتاج صيانة'),
('الإجمالي','=SUM(B2:B3)','للمراجعة فقط','محسوب تلقائيًا'),
])
for r in range(2,4): ws[f'B{r}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('07_IRAP_والطرق'); setup(ws,['القسم','البند','القيمة','الوحدة/النسبة','أين يظهر؟'],[24,34,16,18,46])
append_rows(ws,[
('IRAP','1 STAR',0.02,'نسبة','رسم IRAP'),('IRAP','2 STAR',0.078,'نسبة','رسم IRAP'),('IRAP','3 STAR',0.47,'نسبة','رسم IRAP'),('IRAP','4 STAR',0.39,'نسبة','رسم IRAP'),('IRAP','5 STAR',0.037,'نسبة','رسم IRAP'),
('طرق','طول شبكة الطرق',kv('الطرق','طول شبكة الطرق',1798.1),'KM','مؤشرات تقييم شبكة الطرق'),('طرق','طول آخر/فرعي 1',kv('الطرق','طول آخر/فرعي 1',297),'KM','مؤشرات تقييم شبكة الطرق'),('طرق','طول آخر/فرعي 2',kv('الطرق','طول آخر/فرعي 2',139.5),'KM','مؤشرات تقييم شبكة الطرق'),('طرق','طول آخر/فرعي 3',kv('الطرق','طول آخر/فرعي 3',1462),'KM','مؤشرات تقييم شبكة الطرق'),('طرق','طول آخر/فرعي 4',kv('الطرق','طول آخر/فرعي 4',75.2),'KM','مؤشرات تقييم شبكة الطرق'),
])
for r in range(2,12): ws[f'C{r}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('08_مدارس_منحنيات_سرعة'); setup(ws,['القسم','المطلوب','المنفذ','تم رفع أمر عمل واعتماده','ملاحظة'],[32,16,16,28,50])
append_rows(ws,[
('المدارس',kv('المدارس','المطلوب',0),kv('المدارس','المنفذ',0),kv('المدارس','تم رفع أمر عمل واعتماده',0),''),
('المنحنيات',kv('المنحنيات','المطلوب',0),kv('المنحنيات','المنفذ',0),kv('المنحنيات','تم رفع أمر عمل واعتماده',0),''),
('الطرق التي تتطلب تخفيض سرعة',kv('تخفيض السرعة','المطلوب',0),kv('تخفيض السرعة','المنفذ',0),kv('تخفيض السرعة','تم رفع أمر عمل واعتماده',0),''),
])
for row in range(2,5):
    for col in ['B','C','D']: ws[f'{col}{row}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('09_نصوص_صفحة_2'); setup(ws,['القسم','الترتيب','النص','ملاحظة'],[28,12,90,30])
append_rows(ws,[(r[0],r[1],r[2],r[3]) for r in texts] if texts else [])
for r in range(2, ws.max_row+1): ws[f'C{r}'].fill=PatternFill('solid', fgColor=pale)

ws=wb.create_sheet('10_خريطة_العناصر'); setup(ws,['الشريحة','العنصر في الباوربوينت','التبويب/الخانة','هل يحدث الآن؟'],[12,45,55,18])
append_rows(ws,[(1,'مربع رقم التحويلات الكبير','02_التحويلات!B4 إجمالي التحويلات','نعم'),(1,'رسم التحويلات النظامية/عالية الخطورة','02_التحويلات!B2:B3','نعم'),(1,'رسم حالة وحدات التعداد المروري','06_التعداد_المروري!B2:B3','نعم'),(1,'رسم SOR','03_SOR','نعم'),(1,'رسم الحوادث','04_الحوادث','نعم'),(1,'رسم أوامر العمل','05_أوامر_العمل','نعم'),(1,'رسم IRAP','07_IRAP_والطرق','نعم'),(2,'الإنجازات/التحديات/المستهدفات/التفاصيل','09_نصوص_صفحة_2','نعم')])

for ws in wb.worksheets:
    ws.sheet_view.rightToLeft=True
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    ws.row_dimensions[1].height=28
wb.save(old)
print('WROTE', old)
print('BACKUP', backup, backup.exists())
