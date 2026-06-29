# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys, json, datetime

LOCAL = Path(r'C:/Users/Lenovo-P51/OneDrive/Desktop/بطاقة الاداء/شيت تحديث بطاقة الاداء.xlsx')
GEN = Path(r'C:/Users/Lenovo-P51/Downloads/ppt_achievement_inspect/generate_safety_card.py')
SPREADSHEET_ID = '1VCIPGV7T-hUA4rvmK9kpBbz5Tx6-tPuvhfARsAiSB6U'
SHEET_NAME = '09_نصوص_صفحة_2'

# 1) Patch generator to accept heading-style sheet.
text = GEN.read_text(encoding='utf-8')
old = '''    def rows_by_section(sheet, section):
        if sheet not in wb.sheetnames:
            return []
        out = []
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if text(row[0]) == section and text(row[2]):
                order = row[1] if row[1] is not None else 9999
                out.append((order, text(row[2])))
        return [v for _, v in sorted(out, key=lambda x: x[0])]
'''
new = '''    def rows_by_section(sheet, section):
        if sheet not in wb.sheetnames:
            return []
        aliases = {
            "اهم إنجازات الشهر": {"اهم إنجازات الشهر", "أهم إنجازات الشهر", "إنجازات الشهر", "انجازات الشهر"},
            "اهم تحديات الشهر": {"اهم تحديات الشهر", "أهم تحديات الشهر", "تحديات الشهر"},
            "مستهدفات الشهر القادم": {"مستهدفات الشهر القادم", "مستهدفات الشهر"},
            "تفاصيل إضافية": {"تفاصيل إضافية", "تفاصيل اضافية"},
        }
        wanted = aliases.get(section, {section})
        all_titles = set().union(*aliases.values())
        out = []
        current = None
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            label = text(row[0])
            if label in all_titles:
                current = label
                # Old format: section label repeated on the same row as the text.
                if label in wanted and text(row[2]):
                    order = row[1] if row[1] is not None else 9999
                    out.append((order, text(row[2])))
                continue
            # New clean format: one section heading, then rows beneath it.
            if current in wanted and text(row[2]):
                order = row[1] if row[1] is not None else 9999
                out.append((order, text(row[2])))
        return [v for _, v in sorted(out, key=lambda x: x[0])]
'''
if old in text:
    GEN.write_text(text.replace(old, new), encoding='utf-8')

# 2) Update local Excel tab.
wb = load_workbook(LOCAL)
ws = wb[SHEET_NAME]
ws.delete_rows(1, ws.max_row)
headers = ['القسم', 'الترتيب', 'النص', 'ملاحظة']
ws.append(headers)
sections = [
    ('إنجازات الشهر', 20, [
        'مراجعه اولية لمخططات تحويلات البحر الاحمر للمرحلة الرابعة.',
        'المتابعة والتوجيه لرفع جودة السلامة بالتحويلات.',
        'رصد عدد 40 ملاحظات سلامة واعداد التقارير اللازمة.',
        'تجديد تصاريح العمل المنتهية بالتنسيق مع ادارة الصيانة .',
    ]),
    ('تحديات الشهر', 10, []),
    ('مستهدفات الشهر القادم', 10, [
        'تكثيف الجولات الميدانية على شبكة الطرق',
        'الترتيب مع إدارة السلامة بالفرع لعمل ورشة عمل لتوعية المقاولين بمواصفات تامين مواقع العمل',
        'مراجعة مخططات الاعمال التكميلية لربط طريق المدينة السريع بطريق المعظم',
    ]),
    ('تفاصيل إضافية', 10, [
        '000012020150\t12020150\tTabok to Madina Road\tSensor No. 1',
        '000012030052\t12030052\tDuba to Jordan Road',
    ]),
]
for title, count, existing in sections:
    ws.append([title, '', '', 'عنوان فقط - لا يكتب الإنجاز في هذا السطر'])
    title_row = ws.max_row
    for cell in ws[title_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='407189')
        cell.alignment = Alignment(horizontal='center')
    for i in range(1, count + 1):
        ws.append(['', i, existing[i-1] if i <= len(existing) else '', ''])

for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='A8483D')
    cell.alignment = Alignment(horizontal='center')
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 1).value == '':
        ws.cell(row, 3).fill = PatternFill('solid', fgColor='FFF2CC')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 90
ws.column_dimensions['D'].width = 35
ws.sheet_view.rightToLeft = True
wb.save(LOCAL)

# 3) Push cleaned tab to Google Sheet.
scripts = Path(r'C:/Users/Lenovo-P51/AppData/Local/hermes/skills/productivity/google-workspace/scripts')
sys.path.insert(0, str(scripts))
from google_api import build_service
sheets = build_service('sheets','v4')
values=[]
for row in ws.iter_rows():
    vals=[]
    for cell in row:
        v=cell.value
        if isinstance(v,(datetime.datetime, datetime.date)):
            v=v.strftime('%d/%m/%Y')
        vals.append('' if v is None else v)
    values.append(vals)
sheets.spreadsheets().values().clear(spreadsheetId=SPREADSHEET_ID, range=f"'{SHEET_NAME}'").execute()
sheets.spreadsheets().values().update(spreadsheetId=SPREADSHEET_ID, range=f"'{SHEET_NAME}'!A1", valueInputOption='USER_ENTERED', body={'values': values}).execute()
meta=sheets.spreadsheets().get(spreadsheetId=SPREADSHEET_ID, fields='sheets(properties(title,sheetId))').execute()
sheet_id=[s['properties']['sheetId'] for s in meta['sheets'] if s['properties']['title']==SHEET_NAME][0]
req=[
 {'updateSheetProperties': {'properties': {'sheetId': sheet_id, 'rightToLeft': True, 'gridProperties': {'frozenRowCount': 1}}, 'fields':'rightToLeft,gridProperties.frozenRowCount'}},
 {'repeatCell': {'range': {'sheetId':sheet_id,'startRowIndex':0,'endRowIndex':1}, 'cell': {'userEnteredFormat': {'backgroundColor': {'red':0.66,'green':0.28,'blue':0.24}, 'textFormat': {'foregroundColor': {'red':1,'green':1,'blue':1}, 'bold': True}, 'horizontalAlignment':'CENTER'}}, 'fields':'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)'}},
 {'repeatCell': {'range': {'sheetId':sheet_id,'startRowIndex':1,'endRowIndex':len(values),'startColumnIndex':2,'endColumnIndex':3}, 'cell': {'userEnteredFormat': {'backgroundColor': {'red':1,'green':0.95,'blue':0.80}, 'wrapStrategy': 'WRAP'}}, 'fields':'userEnteredFormat(backgroundColor,wrapStrategy)'}},
]
for idx, row in enumerate(values[1:], start=1):
    if row[0]:
        req.append({'repeatCell': {'range': {'sheetId':sheet_id,'startRowIndex':idx,'endRowIndex':idx+1}, 'cell': {'userEnteredFormat': {'backgroundColor': {'red':0.25,'green':0.44,'blue':0.54}, 'textFormat': {'foregroundColor': {'red':1,'green':1,'blue':1}, 'bold': True}, 'horizontalAlignment':'CENTER'}}, 'fields':'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)'}})
req += [
 {'updateDimensionProperties': {'range': {'sheetId':sheet_id,'dimension':'COLUMNS','startIndex':0,'endIndex':1}, 'properties': {'pixelSize':180}, 'fields':'pixelSize'}},
 {'updateDimensionProperties': {'range': {'sheetId':sheet_id,'dimension':'COLUMNS','startIndex':1,'endIndex':2}, 'properties': {'pixelSize':70}, 'fields':'pixelSize'}},
 {'updateDimensionProperties': {'range': {'sheetId':sheet_id,'dimension':'COLUMNS','startIndex':2,'endIndex':3}, 'properties': {'pixelSize':620}, 'fields':'pixelSize'}},
 {'updateDimensionProperties': {'range': {'sheetId':sheet_id,'dimension':'COLUMNS','startIndex':3,'endIndex':4}, 'properties': {'pixelSize':240}, 'fields':'pixelSize'}},
]
sheets.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body={'requests':req}).execute()
print(json.dumps({'updated_local':str(LOCAL),'updated_google_tab':SHEET_NAME,'rows':len(values)}, ensure_ascii=False, indent=2))
