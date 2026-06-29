# -*- coding: utf-8 -*-
"""
مولّد بطاقة أداء السلامة من Excel إلى PowerPoint.
يحافظ على تصميم ملف PowerPoint الأصلي، ويحدّث النصوص والأرقام وبيانات الرسوم البيانية المخزنة داخل القالب.
"""
from __future__ import annotations

import argparse
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
for k, v in NS.items():
    ET.register_namespace(k, v)


def num(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace("٫", ".").replace(",", "."))
    except Exception:
        return default


def clean_num(v):
    x = num(v, 0)
    return str(int(x)) if abs(x - int(x)) < 1e-9 else str(x)


def text(v):
    return "" if v is None else str(v).strip()


def load_input(xlsx: Path) -> dict:
    wb = load_workbook(xlsx, data_only=False)

    def cell(sheet, ref, default=""):
        if sheet not in wb.sheetnames:
            return default
        v = wb[sheet][ref].value
        return default if v is None else v

    def general(field, default=""):
        sheet = "01_بيانات_عامة" if "01_بيانات_عامة" in wb.sheetnames else "بيانات_عامة"
        if sheet not in wb.sheetnames:
            return default
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if text(row[0]) == field:
                return text(row[1]) or default
        return default

    def rows_by_section(sheet, section):
        if sheet not in wb.sheetnames:
            return []
        aliases = {
            "اهم إنجازات الشهر": {"اهم إنجازات الشهر", "أهم إنجازات الشهر", "إنجازات الشهر", "انجازات الشهر"},
            "اهم تحديات الشهر": {"اهم تحديات الشهر", "أهم تحديات الشهر", "تحديات الشهر"},
            "مستهدفات الشهر القادم": {"مستهدفات الشهر القادم", "مستهدفات الشهر"},
            "تفاصيل إضافية": {"تفاصيل إضافية", "تفاصيل اضافية"},
        }
        wanted = aliases.get(section, {section})
        all_titles = set().union(*aliases.values())
        out = []
        current = None
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            label = text(row[0])
            if label in all_titles:
                current = label
                # Old format: section label repeated on the same row as the text.
                if label in wanted and text(row[2]):
                    order = row[1] if row[1] is not None else 9999
                    out.append((order, text(row[2])))
                continue
            # New clean format: one section heading, then rows beneath it.
            if current in wanted and text(row[2]):
                order = row[1] if row[1] is not None else 9999
                out.append((order, text(row[2])))
        return [v for _, v in sorted(out, key=lambda x: x[0])]

    # new clear workbook values
    normal = num(cell("02_التحويلات", "B2", 0))
    high = num(cell("02_التحويلات", "B3", 0))
    conversion_total = normal + high

    sor_prev = num(cell("03_SOR", "B2", 0))
    sor_month = num(cell("03_SOR", "B3", 0))
    sor_closed = num(cell("03_SOR", "B4", 0))
    sor_order = num(cell("03_SOR", "B5", 0))
    # Prefer an explicit cumulative-total row when present; otherwise preserve the old behavior.
    sor_total_cell = cell("03_SOR", "B6", "")
    sor_total = num(sor_total_cell, sor_prev + sor_month) if sor_total_cell not in ("", None) else sor_prev + sor_month

    # New accident input layout:
    # B2 = previous cumulative accidents, B3 = accidents this month,
    # B4 = cumulative total displayed in the top accident KPI box.
    # B5/B6 keep the chart values for injuries/deaths.
    accident_prev = num(cell("04_الحوادث", "B2", 0))
    accident_month = num(cell("04_الحوادث", "B3", 0))
    accident_total_cell = cell("04_الحوادث", "B4", "")
    accident_total = num(accident_total_cell, accident_prev + accident_month) if accident_total_cell not in ("", None) else accident_prev + accident_month
    injuries = num(cell("04_الحوادث", "B5", cell("04_الحوادث", "B2", 0)))
    deaths = num(cell("04_الحوادث", "B6", cell("04_الحوادث", "B3", 0)))

    work_a = num(cell("05_أوامر_العمل", "B2", 0))
    work_b = num(cell("05_أوامر_العمل", "B3", 0))
    work_c = num(cell("05_أوامر_العمل", "B4", 0))
    work_d = num(cell("05_أوامر_العمل", "B5", 0))
    work_e = num(cell("05_أوامر_العمل", "B6", 0))
    work_total = work_a + work_b + work_c + work_d + work_e

    traffic_working = num(cell("06_التعداد_المروري", "B2", 0))
    traffic_maint = num(cell("06_التعداد_المروري", "B3", 0))

    irap = [num(cell("07_IRAP_والطرق", f"C{i}", 0)) for i in range(2, 7)]
    road_main = cell("07_IRAP_والطرق", "C7", 1798.1)
    road_1 = cell("07_IRAP_والطرق", "C8", 297)
    road_2 = cell("07_IRAP_والطرق", "C9", 139.5)
    road_3 = cell("07_IRAP_والطرق", "C10", 1462)
    road_4 = cell("07_IRAP_والطرق", "C11", 75.2)

    schools = [cell("08_مدارس_منحنيات_سرعة", f"{c}2", 0) for c in "BCD"]
    curves = [cell("08_مدارس_منحنيات_سرعة", f"{c}3", 0) for c in "BCD"]
    speed = [cell("08_مدارس_منحنيات_سرعة", f"{c}4", 0) for c in "BCD"]

    data = {
        "region": general("المنطقة", "تبوك"),
        "month": general("الشهر", "يونيو"),
        "date": general("التاريخ", "03/07/2025"),
        "engineer": general("كبير المهندسين", "م حســـان الــحلــبي"),
        "footer_note": general("ملاحظة أسفل صفحة 1", general("صفحة 1 - ملاحظات سفلية", "التفاصيل الإضافية-يتم ذكرها في الصفحة التالية مع الترميز (*,**,***)")),
        "conversion": {"normal": normal, "high": high, "total": conversion_total},
        "sor": {"prev": sor_prev, "month": sor_month, "closed": sor_closed, "order": sor_order, "total": sor_total},
        "accidents": {"prev": accident_prev, "month": accident_month, "total": accident_total, "injuries": injuries, "deaths": deaths},
        "work_orders": {"A": work_a, "B": work_b, "C": work_c, "D": work_d, "E": work_e, "total": work_total},
        "traffic_counters": {"working": traffic_working, "maintenance": traffic_maint},
        "irap": irap,
        "roads": {"main": road_main, "r1": road_1, "r2": road_2, "r3": road_3, "r4": road_4},
        "schools": schools,
        "curves": curves,
        "speed": speed,
        "achievements": rows_by_section("09_نصوص_صفحة_2", "اهم إنجازات الشهر"),
        "challenges": rows_by_section("09_نصوص_صفحة_2", "اهم تحديات الشهر"),
        "targets": rows_by_section("09_نصوص_صفحة_2", "مستهدفات الشهر القادم"),
        "details": rows_by_section("09_نصوص_صفحة_2", "تفاصيل إضافية"),
    }
    return data


def get_shape_name(sp: ET.Element) -> str:
    nv = sp.find("./p:nvSpPr/p:cNvPr", NS)
    return nv.get("name", "") if nv is not None else ""


def get_texts(sp: ET.Element) -> list[ET.Element]:
    return [t for t in sp.findall(".//a:t", NS)]


def text_content(sp: ET.Element) -> str:
    return "\n".join(t.text or "" for t in get_texts(sp))


def set_shape_text(sp: ET.Element, value: str) -> bool:
    texts = get_texts(sp)
    if not texts:
        return False
    lines = str(value).split("\n") if value is not None else [""]
    texts[0].text = lines[0] if lines else ""
    for i, t in enumerate(texts[1:], 1):
        t.text = lines[i] if i < len(lines) else ""
    return True


def set_by_name(root: ET.Element, name: str, value: str) -> int:
    count = 0
    for sp in root.findall(".//p:sp", NS):
        if get_shape_name(sp) == name:
            if set_shape_text(sp, value):
                count += 1
    return count


def replace_exact_text(root: ET.Element, old: str, new: str) -> int:
    count = 0
    for t in root.findall(".//a:t", NS):
        if (t.text or "").strip() == old:
            t.text = new
            count += 1
    return count


def replace_contains(root: ET.Element, needle: str, new: str) -> int:
    count = 0
    for sp in root.findall(".//p:sp", NS):
        if needle in text_content(sp):
            if set_shape_text(sp, new):
                count += 1
    return count


def update_slide(root: ET.Element, slide_no: int, data: dict) -> dict:
    log = {"slide": slide_no, "updates": []}
    def upd_name(name: str, val):
        c = set_by_name(root, name, str(val))
        log["updates"].append((name, c, str(val)))

    replace_exact_text(root, "بمنطقة (تبوك)", f"بمنطقة ({data['region']})")
    replace_exact_text(root, "03/07/2025", data["date"])
    replace_exact_text(root, "(يونيو)", f"({data['month']})")
    replace_contains(root, "اسم كبير المهندسين:", f"اسم كبير المهندسين: {data['engineer']}")

    if slide_no == 1:
        upd_name("TextBox 58", clean_num(data["conversion"]["total"]))
        # TextBox 2 is the gold number beside work-order review chart in the source template.
        upd_name("TextBox 2", clean_num(data["work_orders"]["total"]))
        r = data["roads"]
        upd_name("TextBox 41", f"{clean_num(r['r1'])}\n KM ")
        upd_name("TextBox 42", f"{clean_num(r['main'])}\n KM")
        upd_name("TextBox 43", f"{clean_num(r['r2'])}\n KM")
        upd_name("TextBox 44", f"{clean_num(r['r3'])}\n KM")
        upd_name("TextBox 45", f"{clean_num(r['r4'])}\n KM")
        for name, val in zip(["TextBox 106","TextBox 105","TextBox 34"], data["schools"]): upd_name(name, clean_num(val))
        for name, val in zip(["TextBox 113","TextBox 114","TextBox 115"], data["curves"]): upd_name(name, clean_num(val))
        for name, val in zip(["TextBox 116","TextBox 117","TextBox 118"], data["speed"]): upd_name(name, clean_num(val))
        # Top accident KPI box in the template; it must be cumulative, not a fixed 100.
        upd_name("TextBox 25", clean_num(data["accidents"]["total"]))
        upd_name("TextBox 36", f"({data['footer_note']})")
    elif slide_no == 2:
        if data["achievements"]: upd_name("TextBox 21", "\n".join(data["achievements"]))
        if data["challenges"]: replace_contains(root, "اهم تحديات الشهر:", "اهم تحديات الشهر:\n" + "\n".join(data["challenges"]))
        if data["targets"]: upd_name("TextBox 2", "\n".join(data["targets"]))
        if data["details"]: upd_name("TextBox 24", "\n".join(data["details"]))
    return log


def set_chart_values(chart_root: ET.Element, values: list[float]) -> int:
    """Update cached numeric values for the first series in a PowerPoint chart."""
    ser = chart_root.find(".//c:ser", NS)
    if ser is None:
        return 0
    pts = ser.findall(".//c:val//c:numCache/c:pt", NS)
    if not pts:
        pts = ser.findall(".//c:val//c:strCache/c:pt", NS)
    updated = 0
    for i, value in enumerate(values):
        if i >= len(pts):
            break
        v = pts[i].find("c:v", NS)
        if v is not None:
            v.text = str(value)
            updated += 1
    return updated


def update_chart(chart_path: str, content: bytes, data: dict) -> tuple[bytes, int]:
    root = ET.fromstring(content)
    mapping = {
        "ppt/charts/chart1.xml": [data["conversion"]["normal"], data["conversion"]["high"]],
        "ppt/charts/chart2.xml": [data["sor"]["total"], data["sor"]["month"], data["sor"]["closed"], data["sor"]["order"]],
        "ppt/charts/chart3.xml": [data["work_orders"]["D"], data["work_orders"]["C"], data["work_orders"]["B"], data["work_orders"]["A"], data["work_orders"]["E"]],
        "ppt/charts/chart4.xml": [data["accidents"]["injuries"], data["accidents"]["deaths"]],
        "ppt/charts/chart5.xml": data["irap"],
        "ppt/charts/chart6.xml": [data["traffic_counters"]["working"], data["traffic_counters"]["maintenance"]],
    }
    if chart_path not in mapping:
        return content, 0
    count = set_chart_values(root, mapping[chart_path])
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), count


def generate(template: Path, data_xlsx: Path, out_pptx: Path) -> list[dict]:
    data = load_input(data_xlsx)
    out_pptx.parent.mkdir(parents=True, exist_ok=True)
    logs = []
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "work.pptx"
        with zipfile.ZipFile(template, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                m = re.match(r"ppt/slides/slide(\d+)\.xml$", item.filename)
                if m:
                    slide_no = int(m.group(1))
                    root = ET.fromstring(content)
                    logs.append(update_slide(root, slide_no, data))
                    content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                # لا نعيد كتابة chart XML مباشرة لأن PowerPoint على Windows قد يرفض بعض ملفات chart المعاد تسلسلها.
                # تحديث الرسوم يتم بعد حفظ الملف عبر PowerPoint COM للحفاظ على سلامة القالب.
                zout.writestr(item, content)
        shutil.move(str(tmp), out_pptx)
    # بعد إنشاء الملف، حدّث الرسوم البيانية عبر PowerPoint نفسه حتى لا تتلف chart XML.
    try:
        chart_logs = update_charts_with_powerpoint(out_pptx, data)
        logs.extend(chart_logs)
    except Exception as exc:
        logs.append({"chart_com_error": str(exc)})
    return logs

def update_charts_with_powerpoint(pptx: Path, data: dict) -> list[dict]:
    """Update PowerPoint embedded chart series using COM on Windows."""
    import pythoncom
    import win32com.client
    updates = {
        "Chart 61": [data["conversion"]["normal"], data["conversion"]["high"]],
        "Chart 79": [data["sor"]["total"], data["sor"]["month"], data["sor"]["closed"], data["sor"]["order"]],
        "Chart 85": [data["work_orders"]["D"], data["work_orders"]["C"], data["work_orders"]["B"], data["work_orders"]["A"], data["work_orders"]["E"]],
        "Chart 39": data["irap"],
        "Chart 47": [data["traffic_counters"]["working"], data["traffic_counters"]["maintenance"]],
    }
    logs=[]; app=None
    pythoncom.CoInitialize()
    try:
        app = win32com.client.Dispatch("PowerPoint.Application")
        app.Visible = 1
        pres = app.Presentations.Open(str(pptx.resolve()), WithWindow=False)
        try:
            for slide in pres.Slides:
                for shp in slide.Shapes:
                    try:
                        if shp.HasChart and shp.Name in updates:
                            shp.Chart.SeriesCollection(1).Values = updates[shp.Name]
                            logs.append({"chart_com": shp.Name, "values": updates[shp.Name]})
                    except Exception:
                        pass
            pres.Save()
        finally:
            pres.Close()
    finally:
        if app:
            app.Quit()
    return logs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True, type=Path)
    ap.add_argument("--data", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()
    logs = generate(args.template, args.data, args.out)
    print(f"تم إنشاء: {args.out}")
    for log in logs:
        if "slide" in log:
            updated = sum(c for _, c, _ in log["updates"])
            print(f"الشريحة {log['slide']}: تحديثات نصية مباشرة = {updated}")
        elif "chart" in log:
            print(f"{log['chart']}: تحديث قيم الرسم = {log['updates']}")
        elif "chart_com" in log:
            print(f"{log['chart_com']}: تحديث الرسم عبر PowerPoint = {log['values']}")
        elif "chart_com_error" in log:
            print(f"تحذير تحديث الرسوم عبر PowerPoint: {log['chart_com_error']}")

if __name__ == "__main__":
    main()
